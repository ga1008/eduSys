# views.py
import logging

import openpyxl
import pandas as pd
from django.contrib.auth import authenticate, login, logout
from django.db import IntegrityError, transaction
from django.db.models import Q, Count
from django.http import HttpResponse
from rest_framework import status
from rest_framework import viewsets
from rest_framework.authtoken.models import Token
from rest_framework.decorators import action
from rest_framework.decorators import api_view, permission_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import AllowAny, IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Class, User, TeacherCourseClass, Course
from .permissions import IsSuperAdmin, IsAdmin, IsTeacher
from .serializers import ClassSerializer, StudentSerializer, TeacherSerializer

logger = logging.getLogger(__name__)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsAdmin])
def only_admin_can_access(request):
    return Response({"detail": "这是只有admin能看到的内容"})


@api_view(["POST"])
@permission_classes([AllowAny])
def login_view(request):
    """处理用户登录，支持邮箱或用户名登录"""
    username_or_email = request.data.get("username")
    password = request.data.get("password")
    requested_role = request.data.get("role")

    # 添加详细日志
    print(f"尝试登录: 用户名/邮箱={username_or_email}, 角色={requested_role}")

    # 将登录过程包含在try-except中，捕获可能的异常
    try:
        user = authenticate(request, username=username_or_email, password=password)

        if user is None:
            return Response({"detail": "用户名/邮箱或密码错误"},
                            status=status.HTTP_401_UNAUTHORIZED)

        # 打印调试信息
        print(f"认证成功: 用户名={user.username}, 角色={user.role}, 请求角色={requested_role}")

        # 如果前端声明了角色，并且与数据库中的user.role不匹配
        if requested_role and user.role != requested_role:
            # 特例：超级管理员可以以管理员身份登录
            if user.role == 'superadmin' and requested_role == 'admin':
                login(request, user)
                return Response({
                    "detail": "登录成功",
                    "role": "superadmin",
                    "username": user.username,
                    "email": user.email,
                }, status=status.HTTP_200_OK)
            else:
                return Response({"detail": f"该账号角色({user.role})与选择的({requested_role})不匹配"},
                                status=status.HTTP_403_FORBIDDEN)

        # 正常登录
        login(request, user)
        return Response({
            "detail": "登录成功",
            "role": user.role,
            "username": user.username,
            "email": user.email,
        }, status=status.HTTP_200_OK)
    except Exception as e:
        print(f"登录异常: {str(e)}")
        return Response({"detail": f"登录过程中发生错误: {str(e)}"},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(["POST"])
@permission_classes([AllowAny])
def logout_view(request):
    # 登出用户，清除会话数据
    logout(request)
    return Response({"detail": "已注销"}, status=status.HTTP_200_OK)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def me_view(request):
    """
    返回当前登录用户的信息, 用于前端刷新时保持同步.
    如果未登录, 默认的IsAuthenticated不通过, 返回403.
    """
    user = request.user  # DRF 根据Session识别出的User
    data = {
        "username": user.username,
        "role": user.role,
        "is_authenticated": user.is_authenticated,
        # 可加更多字段, 如 user.id, email 等
    }
    return Response(data, status=status.HTTP_200_OK)


class TeacherViewSet(viewsets.ModelViewSet):
    queryset = User.objects.filter(role='teacher')
    serializer_class = TeacherSerializer
    permission_classes = [IsAuthenticated, IsAdmin]

    def get_queryset(self):
        qs = super().get_queryset()
        search = self.request.query_params.get('search', '')
        if search:
            qs = qs.filter(
                Q(name__icontains=search) |
                Q(teacher_number__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search)
            )
        return qs


class ClassViewSet(viewsets.ModelViewSet):
    queryset = Class.objects.all()
    serializer_class = ClassSerializer
    permission_classes = [IsAdmin, IsSuperAdmin]

    def get_queryset(self):
        # 使用正确的关联名称计算学生数量
        qs = Class.objects.annotate(
            student_count=Count('students')
        )

        # 如果有搜索参数，添加过滤条件
        search = self.request.query_params.get('search', '')
        if search:
            qs = qs.filter(
                Q(name__icontains=search) |
                Q(year__icontains=search)
            )

        return qs

    @action(detail=True, methods=['get'], url_path='students', url_name='list_students')
    def get_students(self, request, pk=None):
        class_obj = self.get_object()
        students_qs = User.objects.filter(class_enrolled=class_obj, role='student')
        # 若需要处理搜索
        search = request.query_params.get('search', '')
        if search:
            students_qs = students_qs.filter(
                Q(name__icontains=search) | Q(student_number__icontains=search) | Q(email__icontains=search))

        serializer = StudentSerializer(students_qs, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'], url_path='students-count', url_name='students_count')
    def get_students_count(self, request, pk=None):
        """获取班级的学生数量"""
        class_obj = self.get_object()
        students_count = User.objects.filter(class_enrolled=class_obj, role='student').count()
        return Response({"students_count": students_count})

    @action(detail=True, methods=['get'])
    def students(self, request, pk=None):
        """获取指定班级的所有学生"""
        class_obj = self.get_object()
        students = User.objects.filter(class_enrolled=class_obj, role='student')

        # 支持搜索功能
        search = request.query_params.get('search', '')
        if search:
            students = students.filter(
                Q(name__icontains=search) |
                Q(student_number__icontains=search) |
                Q(email__icontains=search)
            )

        page = self.paginate_queryset(students)
        if page is not None:
            serializer = StudentSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = StudentSerializer(students, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='import-students', parser_classes=[MultiPartParser, FormParser])
    def import_students(self, request, pk=None):
        """批量导入学生到指定班级"""
        gender_map = {'男': 'M', '女': 'F', '其他': 'O'}

        # 获取班级对象
        class_obj = self.get_object()

        # 获取上传的文件
        file = request.FILES.get('file')
        if not file:
            return Response({"detail": "未提供文件"}, status=status.HTTP_400_BAD_REQUEST)

        # 导入结果统计
        stats = {
            "success": 0,
            "failed": 0,
            "errors": []
        }

        # 预处理数据列表
        students_to_create = []

        try:
            # 使用pandas统一处理CSV和Excel文件
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
            elif file.name.endswith('.xlsx') or file.name.endswith('.xls'):
                df = pd.read_excel(file)
            else:
                return Response({"detail": "不支持的文件格式"}, status=status.HTTP_400_BAD_REQUEST)

            # 规范化列名
            df.columns = [col.lower().strip() for col in df.columns]

            # 映射常见的列名
            column_mapping = {
                'student_number': ['student_number', 'student_id', '学号'],
                'name': ['name', '姓名'],
                'email': ['email', '邮箱'],
                'gender': ['gender', '性别'],
                'phone': ['phone', '手机', '电话'],
                'qq': ['qq'],
                'password': ['password', '初始密码']
            }

            # 创建标准化的DataFrame
            processed_data = {}
            for standard, alternatives in column_mapping.items():
                for alt in alternatives:
                    if alt in df.columns:
                        processed_data[standard] = df[alt]
                        break

            # 转换为标准DataFrame
            std_df = pd.DataFrame(processed_data)

            # 验证必填字段
            required_fields = ['student_number', 'name', 'email', 'password']
            for field in required_fields:
                if field not in std_df.columns:
                    return Response({
                        "detail": f"文件缺少必填字段: {field}"
                    }, status=status.HTTP_400_BAD_REQUEST)

            # 处理每一行数据
            for _, row in std_df.iterrows():
                try:
                    # 基本数据验证
                    student_number = str(row.get('student_number')).strip()
                    name = str(row.get('name')).strip()
                    email = str(row.get('email')).strip()
                    gender = row.get('gender', '').strip() if pd.notna(row.get('gender')) else ''
                    phone = str(row.get('phone', '')).strip() if pd.notna(row.get('phone')) else ''
                    qq = str(row.get('qq', '')).strip() if pd.notna(row.get('qq')) else ''
                    password = str(row.get('password')).strip()

                    # 验证必填项
                    if not all([student_number, name, email, password]):
                        stats["failed"] += 1
                        stats["errors"].append(f"学号{student_number}：必填字段不完整")
                        continue

                    # 邮箱格式验证
                    if '@' not in email:
                        stats["failed"] += 1
                        stats["errors"].append(f"学号{student_number}：邮箱格式不正确")
                        continue

                    # 转换性别格式
                    if gender:
                        gender = gender_map.get(gender, 'O')
                    else:
                        gender = 'O'

                    # 创建用户对象
                    user = User(
                        username=student_number,
                        student_number=student_number,
                        name=name,
                        email=email,
                        gender=gender,
                        phone=phone,
                        qq=qq,
                        role='student',
                        class_enrolled=class_obj
                    )

                    # 设置密码
                    user.set_password(password)

                    # 添加到待创建列表
                    students_to_create.append(user)
                    stats["success"] += 1

                except Exception as e:
                    stats["failed"] += 1
                    stats["errors"].append(f"学号{student_number if 'student_number' in locals() else '未知'}：{str(e)}")

            # 使用事务进行批量创建
            with transaction.atomic():
                batch_size = 100  # 每批处理的记录数
                for i in range(0, len(students_to_create), batch_size):
                    batch = students_to_create[i:i + batch_size]
                    try:
                        # 先检查用户名是否已存在
                        usernames = [u.username for u in batch]
                        existing = User.objects.filter(username__in=usernames).values_list('username', flat=True)

                        if existing:
                            for username in existing:
                                stats["success"] -= 1
                                stats["failed"] += 1
                                stats["errors"].append(f"学号{username}：用户已存在")

                            # 过滤掉已存在的用户
                            batch = [u for u in batch if u.username not in existing]

                        # 批量创建用户
                        if batch:
                            User.objects.bulk_create(batch)

                    except IntegrityError as e:
                        stats["success"] -= len(batch)
                        stats["failed"] += len(batch)
                        stats["errors"].append(f"批量导入失败：{str(e)}")

            # 返回导入结果统计
            return Response({
                "detail": f"导入完成: 成功 {stats['success']} 条, 失败 {stats['failed']} 条",
                "stats": stats
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                "detail": f"导入过程发生错误: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['delete'])
    def students(self, request, pk=None):
        """清空班级学生"""
        class_obj = self.get_object()
        students = User.objects.filter(class_enrolled=class_obj, role='student')
        count = students.count()
        students.delete()

        return Response({
            "detail": f"已删除 {count} 名学生"
        }, status=status.HTTP_200_OK)


class StudentViewSet(viewsets.ModelViewSet):
    queryset = User.objects.filter(role='student')
    serializer_class = StudentSerializer
    permission_classes = [IsSuperAdmin, IsTeacher]  # 超级管理员权限

    def get_queryset(self):
        """如果提供了 class_id 参数，则过滤只返回该班级的学生"""
        qs = super().get_queryset()
        class_id = self.request.query_params.get('class_id')
        if class_id:
            qs = qs.filter(class_enrolled_id=class_id)
        return qs

    @action(detail=False, methods=['get'])
    def download_template(self, request, pk=None):
        """下载学生导入模板"""
        format_type = request.query_params.get('format', 'xlsx')

        if format_type == 'xlsx':
            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            ws = wb.active

            # 写入表头
            headers = ['学号', '姓名', 'email', '初始密码', '性别', '手机', 'QQ']
            ws.append(headers)

            # 写入示例数据
            ws.append(['10001', '张三', 'zhangsan@example.com', '123456', 'M', '13800138000', '10001'])
            ws.append(['10002', '李四', 'lisi@example.com', '123456', 'F', '13800138001', '10002'])

            # 设置响应头
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=学生导入模板.xlsx'

            # 保存到响应
            wb.save(response)
            return response


# backend/education/views.py
class StudentBulkUpdateView(APIView):
    permission_classes = [IsSuperAdmin]

    def put(self, request, class_id):
        """批量更新学生信息"""
        students_data = request.data
        updated = []
        errors = []

        for student in students_data:
            try:
                user = User.objects.get(id=student['id'])
                serializer = StudentSerializer(user, data=student, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    updated.append(serializer.data)
                else:
                    errors.append({
                        'id': student['id'],
                        'errors': serializer.errors
                    })
            except User.DoesNotExist:
                errors.append({
                    'id': student['id'],
                    'errors': '学生不存在'
                })

        return Response({
            'updated': updated,
            'errors': errors
        })

    def delete(self, request, class_id):
        """清空班级学生"""
        try:
            class_obj = Class.objects.get(id=class_id)
            class_obj.students.all().delete()
            return Response({'detail': '已清空该班级所有学生'})
        except Class.DoesNotExist:
            return Response({'detail': '班级不存在'}, status=404)


class ImportStudentsView(APIView):
    permission_classes = [IsSuperAdmin]
    parser_classes = [MultiPartParser]  # 支持multipart文件上传

    def post(self, request, pk, format=None):
        """批量导入学生到指定班级 (班级ID由URL中的pk指定)"""
        gender_map = {
            '男': 'M',
            '女': 'F',
            '其他': 'O'
        }
        # 校验班级存在
        try:
            class_obj = Class.objects.get(pk=pk)
        except Class.DoesNotExist:
            return Response({"detail": "班级不存在"}, status=status.HTTP_404_NOT_FOUND)
        # 获取上传的文件
        file = request.FILES.get('file')
        if not file:
            return Response({"detail": "未提供文件"}, status=status.HTTP_400_BAD_REQUEST)
        new_students = []  # 保存成功创建的学生用户
        # 根据文件扩展名分别处理 CSV 和 Excel
        if file.name.endswith('.csv'):
            import csv, io
            # 读取CSV文件内容并解析
            data = file.read().decode('utf-8')
            f = io.StringIO(data)
            reader = csv.DictReader(f)
            for row in reader:
                # 假设CSV包含以下列（可根据实际文件调整列名）:
                # student_number, name, email, password (初始密码)
                student_number = row.get('student_number') or row.get('student_id') or row.get('学号')
                name = row.get('name') or row.get('姓名')
                email = row.get('email')
                gender = row.get('gender') or row.get('性别')
                if gender:
                    gender = gender_map.get(gender, 'O')  # 默认值为'O'

                password = row.get('password') or row.get('初始密码')
                # 基本校验，跳过缺少必要字段的行
                if not (student_number and name and email and password):
                    continue
                # 创建User对象并设置字段
                user = User(
                    username=str(student_number),  # 用学号作为用户名
                    student_number=str(student_number),
                    name=str(name),
                    email=str(email),
                    gender=str(gender),
                    role='student',
                    class_enrolled=class_obj
                )
                user.set_password(str(password))  # 设置初始密码（哈希保存）
                try:
                    user.save()
                except Exception as e:
                    # 如遇到唯一约束冲突等错误，可以记录或跳过
                    continue
                new_students.append(user)
        elif file.name.endswith('.xls') or file.name.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            # 读取表头
            headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                row_data = dict(zip(headers, row))
                student_number = row_data.get('student_number') or row_data.get('student_id') or row_data.get('学号')
                name = row_data.get('name') or row_data.get('姓名')
                gender = row_data.get('gender') or row_data.get('性别')
                if gender:
                    gender = gender_map.get(gender, 'O')  # 默认值为'O'

                email = row_data.get('email')
                password = row_data.get('password') or row_data.get('初始密码')
                if not (student_number and name and email and password):
                    continue
                user = User(
                    username=str(student_number),
                    student_number=str(student_number),
                    name=str(name),
                    gender=str(gender),
                    email=str(email),
                    role='student',
                    class_enrolled=class_obj
                )
                user.set_password(str(password))
                try:
                    user.save()
                except Exception as e:
                    continue
                new_students.append(user)
        else:
            return Response({"detail": "不支持的文件格式"}, status=status.HTTP_400_BAD_REQUEST)
        # （可选）可将数据库操作放入事务，或使用bulk_create提高性能&#8203;:contentReference[oaicite:7]{index=7}
        # 将新创建的学生序列化返回
        serializer = StudentSerializer(new_students, many=True)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
